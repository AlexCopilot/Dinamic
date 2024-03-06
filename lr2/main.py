import telebot
import openpyxl
from openpyxl import load_workbook



token = '6361049571:AAG3NWySibKBqDTANNqF9FYv-Wri2xcD-jE'
bot = telebot.TeleBot(token)
START = ("""Добро пожаловать в телеграмм-бота, здесь Вы можете узнать успеваемость своего ребенка.

Чтобы узнать успеваемость ребенка, напишите его ФИО полностью""")
HELP = "Чтобы узнать успеваемость ребенка, напишите его ФИО полностью"


@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, START)

@bot.message_handler(commands=['help'])
def help(message):
    bot.send_message(message.chat.id, HELP)

@bot.message_handler(content_types='text')
def perfomance(message):
    bot.send_message(message.chat.id, query_handler(message))

def query_handler(message):
    student = message.text
    file_path = 'Perfomance.xlsx'
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        bot.send_message(message.chat.id, "Файл с успеваемостью не найден.")
        return

    message_text = "Оценки ученика " + student + ":\n"
    marks = ''
    try:
        ws = wb['Русский язык']
        for row in ws.iter_rows(min_row=2, max_col=2):
            if row[0].value == student:
                marks += str(row[1].value) + ' '
            elif row[0].value == '':
                break
        message_text += "Русский язык: " + marks + "\n"
        marks = ''
    except:
        pass

    try:
        ws = wb['Математика']
        for row in ws.iter_rows(min_row=2, max_col=2):
            if row[0].value == student:
                marks += str(row[1].value) + ' '
            elif row[0].value == '':
                break
        message_text += "Математика: " + marks + "\n"
        marks = ''
    except:
        pass

    try:
        ws = wb['Литература']
        for row in ws.iter_rows(min_row=2, max_col=2):
            if row[0].value == student:
                marks += str(row[1].value) + ' '
            elif row[0].value == '':
                break
        message_text += "Литература: " + marks + "\n"
        marks = ''
    except:
        pass


    return message_text

bot.polling(none_stop=True)
