import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
url = 'https://omsk.cian.ru/recommendations/'
r = requests.get(url)
soup = BeautifulSoup(r.text, 'html.parser')
blocks = soup.findAll('div', class_='_4d935d0799--container--ktXmQ _4d935d0799--offer--bMJvM')
file_path = 'Objects.xlsx'
try:
    wb = load_workbook(file_path)
    ws = wb['data']
except:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'data'

ws['A1'] = 'Ссылка'
ws['B1'] = 'Цена'
ws['C1'] = 'Тип объекта'
ws['D1'] = 'Площадь'
ws['E1'] = 'Адресс'
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 9
ws.column_dimensions['E'].width = 100
for block in blocks:
    link = 'https://omsk.cian.ru' + block.find('a', class_='_4d935d0799--link--acwir').get('href')
    price = block.find('div', class_='_4d935d0799--price--hSzzN').find('span',class_='_4d935d0799--color_black_100--Ephi7 _4d935d0799--lineHeight_7u--jtkAy _4d935d0799--fontWeight_bold--BbhnX _4d935d0799--fontSize_22px--sFuaL _4d935d0799--display_block--KYb25 _4d935d0799--text--e4SBY').text
    object_type = (block.find('a', class_='_4d935d0799--link--ZMT_z').text).split('•')[0].replace('кв.','квартира').replace('апарт.', 'апартаменты')
    object_area = (block.find('a', class_='_4d935d0799--link--ZMT_z').text).split('•')[1].replace('"\"xa0', '').replace(' ', '')
    address = block.find('span',class_='_4d935d0799--color_black_60--wABx7 _4d935d0799--lineHeight_5u--e6Sug _4d935d0799--fontWeight_normal--JEG_c _4d935d0799--fontSize_14px--reQMB _4d935d0799--display_block--KYb25 _4d935d0799--text--e4SBY _4d935d0799--text_letterSpacing__0--cQxU5').text
    ws.append([link, price, object_type, object_area, address])
wb.save(file_path)
wb.close()
